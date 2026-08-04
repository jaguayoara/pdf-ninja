"""
Modulo de conversion de PDF a otros formatos y viceversa.
- PDF -> Word (.docx)
- PDF -> Excel (.xlsx) extrayendo tablas
- PDF -> Imagen (PNG/JPG por pagina o zip)
- Imagen(es) -> PDF
- PDF -> Texto plano
- PDF -> PDF/A (basico via reescritura)
"""
from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import List, Optional, Tuple

import fitz  # PyMuPDF
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from pdf2docx import Converter

from . import utils


# --- PDF -> Word ---------------------------------------------------------

def pdf_to_word(pdf_path: Path, out_path: Path) -> Path:
    """Convierte PDF a DOCX preservando layout basico (texto, tablas, imagenes)."""
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    cv = Converter(str(pdf_path))
    try:
        cv.convert(str(out_path))
    finally:
        cv.close()
    if not out_path.exists() or out_path.stat().st_size == 0:
        raise RuntimeError("Conversion PDF->Word fallo: archivo vacio")
    return out_path


# --- PDF -> Excel --------------------------------------------------------

def pdf_to_excel(pdf_path: Path, out_path: Path) -> Path:
    """
    Extrae tablas de cada pagina del PDF a hojas de un .xlsx.
    Si una pagina no tiene tablas, agrega una hoja 'Texto_pagina_N' con el texto.
    """
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Quitar la hoja por defecto
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C7BE5")
    wrap = Alignment(wrap_text=True, vertical="top")

    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables() or []
            if tables:
                for t_idx, table in enumerate(tables, start=1):
                    sheet_name = f"P{i}_Tabla{t_idx}" if len(tables) > 1 else f"Pagina_{i}"
                    # Excel limita nombres a 31 chars
                    sheet_name = sheet_name[:31]
                    if sheet_name in wb.sheetnames:
                        sheet_name = f"{sheet_name}_{t_idx}"
                    ws = wb.create_sheet(sheet_name)
                    if not table:
                        continue
                    for r_idx, row in enumerate(table, start=1):
                        for c_idx, cell in enumerate(row, start=1):
                            cell_value = "" if cell is None else str(cell)
                            cell_obj = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                            cell_obj.alignment = wrap
                            if r_idx == 1:
                                cell_obj.font = header_font
                                cell_obj.fill = header_fill
                    # Auto-ancho basico
                    for c_idx in range(1, (max((len(r) for r in table), default=1)) + 1):
                        max_len = 0
                        for r_idx, row in enumerate(table, start=1):
                            v = row[c_idx - 1] if c_idx - 1 < len(row) and row[c_idx - 1] is not None else ""
                            max_len = max(max_len, min(len(str(v)), 60))
                        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len + 2, 10), 60)
                    ws.freeze_panes = "A2"
            else:
                # Fallback: texto plano
                text = page.extract_text() or ""
                sheet_name = f"Texto_p{i}"[:31]
                if sheet_name in wb.sheetnames:
                    sheet_name = f"{sheet_name}_x"
                ws = wb.create_sheet(sheet_name)
                for r_idx, line in enumerate(text.splitlines() or [""], start=1):
                    ws.cell(row=r_idx, column=1, value=line).alignment = wrap
                ws.column_dimensions["A"].width = 120

    if not wb.sheetnames:
        wb.create_sheet("Vacio")

    wb.save(str(out_path))
    if not out_path.exists() or out_path.stat().st_size == 0:
        raise RuntimeError("Conversion PDF->Excel fallo: archivo vacio")
    return out_path


# --- PDF -> Imagen -------------------------------------------------------

def pdf_to_images(pdf_path: Path, out_dir: Path, fmt: str = "png", dpi: int = 150) -> List[Path]:
    """
    Convierte cada pagina a una imagen (PNG/JPG).
    Devuelve la lista de archivos creados.
    """
    utils.ensure_pdf(pdf_path)
    out_dir.mkdir(parents=True, exist_ok=True)
    fmt = fmt.lower()
    if fmt not in {"png", "jpg", "jpeg"}:
        fmt = "png"
    if fmt == "jpeg":
        fmt = "jpg"

    results: List[Path] = []
    doc = fitz.open(str(pdf_path))
    try:
        zoom = dpi / 72.0
        mat = fitz.Matrix(zoom, zoom)
        for i, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=mat, alpha=False)
            out_file = out_dir / f"pagina_{i:03d}.{fmt}"
            if fmt == "png":
                pix.save(str(out_file))
            else:
                # Convertir a PIL para JPG
                img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
                img.save(str(out_file), "JPEG", quality=92, optimize=True)
            results.append(out_file)
    finally:
        doc.close()
    return results


def pdf_to_images_zip(pdf_path: Path, out_zip: Path, fmt: str = "png", dpi: int = 150) -> Path:
    """Convierte PDF a imagenes y las empaqueta en un .zip."""
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        tmp_dir = Path(td)
        imgs = pdf_to_images(pdf_path, tmp_dir, fmt=fmt, dpi=dpi)
        out_zip.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(str(out_zip), "w", zipfile.ZIP_DEFLATED) as zf:
            for img in imgs:
                zf.write(str(img), arcname=img.name)
    return out_zip


# --- Imagen -> PDF -------------------------------------------------------

def images_to_pdf(image_paths: List[Path], out_path: Path) -> Path:
    """Une varias imagenes en un PDF, una imagen por pagina."""
    if not image_paths:
        raise ValueError("No se proporcionaron imagenes")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    pil_images: List[Image.Image] = []
    try:
        for p in image_paths:
            if not p.exists():
                raise FileNotFoundError(f"No existe: {p}")
            img = Image.open(str(p))
            if img.mode in ("RGBA", "LA", "P"):
                # PDF no soporta alpha bien: convertir a RGB
                bg = Image.new("RGB", img.size, (255, 255, 255))
                if img.mode == "P":
                    img = img.convert("RGBA")
                bg.paste(img, mask=img.split()[-1] if img.mode in ("RGBA", "LA") else None)
                img = bg
            elif img.mode != "RGB":
                img = img.convert("RGB")
            pil_images.append(img)

        first, rest = pil_images[0], pil_images[1:]
        first.save(
            str(out_path),
            "PDF",
            save_all=True,
            append_images=rest,
            resolution=150.0,
        )
    finally:
        for im in pil_images:
            try:
                im.close()
            except Exception:
                pass

    if not out_path.exists() or out_path.stat().st_size == 0:
        raise RuntimeError("Conversion imagen->PDF fallo: archivo vacio")
    return out_path


# --- PDF -> Texto --------------------------------------------------------

def pdf_to_text(pdf_path: Path, out_path: Path) -> Path:
    """Extrae texto plano del PDF."""
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    chunks: List[str] = []
    doc = fitz.open(str(pdf_path))
    try:
        for i, page in enumerate(doc, start=1):
            text = page.get_text("text") or ""
            chunks.append(f"--- Pagina {i} ---\n{text}\n")
    finally:
        doc.close()
    out_path.write_text("\n".join(chunks), encoding="utf-8")
    return out_path


# --- PDF -> PDF/A (basico, reescritura con fonts embebidas) -------------

def pdf_to_pdfa(pdf_path: Path, out_path: Path) -> Path:
    """
    Re-escribe el PDF intentando embeber fonts y normalizar.
    No es un PDF/A-1 estricto certificado, pero es compatible con muchos visores.
    Para PDF/A estricto se recomienda usar tools externas.
    """
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(str(pdf_path))
    try:
        # Forzar embedding de fuentes
        for page in doc:
            for f in page.get_fonts(full=True):
                # xref, ext, type, basefont, name, encoding, referencer
                xref = f[0]
                try:
                    page.insert_font(xref)  # noop, fuerza registro
                except Exception:
                    pass
        # Re-escribir con garbage collection
        doc.save(
            str(out_path),
            garbage=4,
            deflate=True,
            clean=True,
            ascii=False,
        )
    finally:
        doc.close()
    return out_path
