"""
Modulo de manipulacion de PDFs: dividir, fusionar, rotar, comprimir,
organizar, marca de agua, numeros de pagina, proteger/desproteger.
Tambien expone extractores de metadatos para PDF, Word, Excel,
PowerPoint, imagenes y texto.
"""
from __future__ import annotations

import io
import math
import re
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import fitz  # PyMuPDF
import pikepdf
from PIL import Image

from . import utils


# --- Fusionar ------------------------------------------------------------

def merge_pdfs(pdf_paths: Sequence[Path], out_path: Path) -> Path:
    """Fusiona varios PDFs en uno solo, en el orden dado."""
    if not pdf_paths:
        raise ValueError("No se proporcionaron PDFs para fusionar")
    for p in pdf_paths:
        utils.ensure_pdf(p)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    result = fitz.open()
    try:
        for p in pdf_paths:
            with fitz.open(str(p)) as src:
                result.insert_pdf(src)
        result.save(str(out_path), garbage=4, deflate=True)
    finally:
        result.close()
    return out_path


# --- Dividir -------------------------------------------------------------

def split_pdf(pdf_path: Path, out_dir: Path, groups: List[List[int]]) -> List[Path]:
    """
    Divide el PDF segun `groups` (lista de listas de paginas 1-based).
    Cada grupo produce un PDF independiente. Los grupos pueden superponerse.
    """
    utils.ensure_pdf(pdf_path)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_dir = out_dir.resolve()

    # Validar y deduplicar grupos
    cleaned: List[List[int]] = []
    for g in groups:
        if not g:
            continue
        seen = set()
        ug = []
        for n in g:
            if n in seen:
                continue
            seen.add(n)
            ug.append(n)
        cleaned.append(sorted(ug))
    if not cleaned:
        raise ValueError("Los rangos de paginas no producen ningun grupo")

    src = fitz.open(str(pdf_path))
    outputs: List[Path] = []
    try:
        total = src.page_count
        for idx, group in enumerate(cleaned, start=1):
            out_doc = fitz.open()
            for n in group:
                if 1 <= n <= total:
                    out_doc.insert_pdf(src, from_page=n - 1, to_page=n - 1)
            if out_doc.page_count == 0:
                out_doc.close()
                continue
            out_file = out_dir / f"{pdf_path.stem}_parte_{idx}.pdf"
            out_file = utils.ensure_within_dir(out_file, out_dir)
            out_doc.save(str(out_file), garbage=4, deflate=True)
            out_doc.close()
            outputs.append(out_file)
    finally:
        src.close()

    if not outputs:
        raise RuntimeError("La division no produjo archivos validos")
    return outputs


def split_pdf_zip(pdf_path: Path, out_zip: Path, groups: List[List[int]]) -> Path:
    """Igual a split_pdf pero devuelve un .zip con todas las partes."""
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        files = split_pdf(pdf_path, tmp, groups)
        with __import__("zipfile").ZipFile(str(out_zip), "w", __import__("zipfile").ZIP_DEFLATED) as zf:
            for f in files:
                zf.write(str(f), arcname=f.name)
    return out_zip


# --- Organizar (reordenar / eliminar) -----------------------------------

def organize_pdf(pdf_path: Path, out_path: Path, order: List[int]) -> Path:
    """
    Reordena y/o elimina paginas.
    `order` es una lista de indices 1-based en el orden deseado.
    """
    utils.ensure_pdf(pdf_path)
    if not order:
        raise ValueError("Debe especificar al menos una pagina")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    src = fitz.open(str(pdf_path))
    try:
        total = src.page_count
        for n in order:
            if not (1 <= n <= total):
                raise ValueError(f"Pagina fuera de rango: {n}")
        out_doc = fitz.open()
        for n in order:
            out_doc.insert_pdf(src, from_page=n - 1, to_page=n - 1)
        out_doc.save(str(out_path), garbage=4, deflate=True)
        out_doc.close()
    finally:
        src.close()
    return out_path


# --- Rotar ---------------------------------------------------------------

def rotate_pdf(pdf_path: Path, out_path: Path, angle: int, pages: Optional[List[int]] = None) -> Path:
    """
    Rota paginas. `angle` debe ser multiplo de 90 (90, 180, 270, -90, etc).
    Si `pages` es None, rota todas.
    """
    if angle % 90 != 0:
        raise ValueError("El angulo debe ser multiplo de 90 grados")
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    src = fitz.open(str(pdf_path))
    try:
        if pages is None:
            pages = list(range(1, src.page_count + 1))
        for n in pages:
            if not (1 <= n <= src.page_count):
                raise ValueError(f"Pagina fuera de rango: {n}")
            page = src[n - 1]
            page.set_rotation((page.rotation + angle) % 360)
        src.save(str(out_path), garbage=4, deflate=True)
    finally:
        src.close()
    return out_path


# --- Comprimir -----------------------------------------------------------

def compress_pdf(pdf_path: Path, out_path: Path, quality: str = "medium") -> Path:
    """
    Reduce el tamano del PDF.
    quality: 'low' | 'medium' | 'high'
      low    -> rasteriza paginas a ~72 dpi JPG baja calidad
      medium -> rasteriza a ~120 dpi JPG calidad media (default)
      high   -> rasteriza a ~150 dpi JPG alta calidad
    NOTA: la rasterizacion convierte el PDF en imagenes, por lo que el texto
    deja de ser seleccionable. Para compresion sin perder texto, prueba primero
    'optimize' (aun no implementado para PDFs muy grandes).
    """
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    settings = {
        "low":    {"dpi": 72,  "jpg_quality": 50},
        "medium": {"dpi": 120, "jpg_quality": 72},
        "high":   {"dpi": 150, "jpg_quality": 85},
    }
    s = settings.get(quality, settings["medium"])

    src = fitz.open(str(pdf_path))
    try:
        out_doc = fitz.open()
        zoom = s["dpi"] / 72.0
        mat = fitz.Matrix(zoom, zoom)
        for page in src:
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            buf = io.BytesIO()
            img.save(buf, "JPEG", quality=s["jpg_quality"], optimize=True)
            buf.seek(0)
            # Crear nueva pagina con mismo tamano en puntos
            rect = fitz.Rect(0, 0, page.rect.width, page.rect.height)
            new_page = out_doc.new_page(width=rect.width, height=rect.height)
            new_page.insert_image(rect, stream=buf.getvalue())
        out_doc.save(str(out_path), garbage=4, deflate=True)
        out_doc.close()
    finally:
        src.close()
    return out_path


# --- Marca de agua -------------------------------------------------------

def add_watermark(
    pdf_path: Path,
    out_path: Path,
    text: str,
    opacity: float = 0.3,
    fontsize: int = 48,
    rotation: int = 45,
    color: Tuple[float, float, float] = (0.5, 0.5, 0.5),
    position: str = "center",  # 'center' | 'tile' | 'top' | 'bottom'
    pages: Optional[List[int]] = None,
) -> Path:
    """Anade marca de agua de texto al PDF. Soporta rotacion arbitraria."""
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not text.strip():
        raise ValueError("El texto de la marca de agua no puede estar vacio")
    if not (0.05 <= opacity <= 1.0):
        raise ValueError("Opacidad debe estar entre 0.05 y 1.0")
    # Mezclar opacidad con el color
    bg = (1.0, 1.0, 1.0)
    blended = tuple(color[i] * opacity + bg[i] * (1 - opacity) for i in range(3))
    src = fitz.open(str(pdf_path))
    try:
        if pages is None:
            pages = list(range(1, src.page_count + 1))
        font = fitz.Font("helv")
        for n in pages:
            if not (1 <= n <= src.page_count):
                raise ValueError(f"Pagina fuera de rango: {n}")
            page = src[n - 1]
            rect = page.rect

            def draw_at(cx: float, cy: float) -> None:
                # Calcular ancho del texto y colocarlo centrado horizontalmente
                try:
                    text_w = fitz.get_text_length(text, fontname="helv", fontsize=fontsize)
                except Exception:
                    text_w = fontsize * 0.55 * max(len(text), 1)
                # Posicion: baseline a la altura de cy, x = cx - text_w/2
                pos = (cx - text_w / 2, cy + fontsize * 0.35)
                tw = fitz.TextWriter(page.rect)
                tw.append(pos, text, font=font, fontsize=fontsize)
                # Rotar el texto alrededor de (cx, cy) y aplicar color
                # M = T(cx,cy) * R(theta) * T(-cx,-cy)
                rot_mat = fitz.Matrix(1, 0, 0, 1, -cx, -cy)
                rot_mat = rot_mat.prerotate(rotation)
                rot_mat = rot_mat.pretranslate(cx, cy)
                tw.write_text(page, color=blended, matrix=rot_mat)

            if position == "tile":
                step_x = max(150, fontsize * 6)
                step_y = max(150, fontsize * 6)
                y = 0
                while y < rect.height:
                    x = 0
                    while x < rect.width:
                        draw_at(x + step_x / 2, y + step_y / 2)
                        x += step_x
                    y += step_y
            elif position == "top":
                draw_at(rect.width / 2, rect.height * 0.15)
            elif position == "bottom":
                draw_at(rect.width / 2, rect.height * 0.85)
            else:  # center
                draw_at(rect.width / 2, rect.height / 2)

        src.save(str(out_path), garbage=4, deflate=True, encryption=fitz.PDF_ENCRYPT_KEEP)
    finally:
        src.close()
    return out_path


# --- Numeros de pagina ---------------------------------------------------

def add_page_numbers(
    pdf_path: Path,
    out_path: Path,
    position: str = "bottom-center",  # bottom-center | bottom-right | top-center | top-right
    start: int = 1,
    fontsize: int = 11,
    color: Tuple[float, float, float] = (0, 0, 0),
    prefix: str = "",
    suffix: str = "",
) -> Path:
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    src = fitz.open(str(pdf_path))
    try:
        total = src.page_count
        for i, page in enumerate(src, start=1):
            rect = page.rect
            n = start + (i - 1)
            label = f"{prefix}{n}{suffix}"
            margin = 24
            if position.startswith("bottom"):
                y = rect.height - margin
            else:
                y = margin + fontsize
            if position.endswith("right"):
                box = fitz.Rect(rect.width - margin - 80, y - fontsize, rect.width - margin, y + 4)
                align = 2  # right
            elif position.endswith("left"):
                box = fitz.Rect(margin, y - fontsize, margin + 80, y + 4)
                align = 0  # left
            else:  # center
                box = fitz.Rect(rect.width / 2 - 40, y - fontsize, rect.width / 2 + 40, y + 4)
                align = 1
            page.insert_textbox(box, label, fontsize=fontsize, color=color, align=align, overlay=True)
        src.save(str(out_path), garbage=4, deflate=True)
    finally:
        src.close()
    return out_path


# --- Proteger (password) ------------------------------------------------

def protect_pdf(pdf_path: Path, out_path: Path, user_pwd: str, owner_pwd: Optional[str] = None) -> Path:
    """Cifra el PDF con password (AES-256)."""
    if not user_pwd:
        raise ValueError("La contrasena no puede estar vacia")
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    src = pikepdf.open(str(pdf_path))
    try:
        owner = owner_pwd or user_pwd
        # AES-256 con R6
        src.save(
            str(out_path),
            encryption=pikepdf.Encryption(
                user=user_pwd,
                owner=owner,
                aes=True,
                R=6,
            ),
        )
    finally:
        src.close()
    return out_path


# --- Desproteger --------------------------------------------------------

def unlock_pdf(pdf_path: Path, out_path: Path, password: str) -> Path:
    """Quita la contrasena del PDF."""
    if not password:
        raise ValueError("Debes proporcionar la contrasena")
    utils.ensure_pdf(pdf_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        src = pikepdf.open(str(pdf_path), password=password)
    except pikepdf.PasswordError as e:
        raise ValueError("Contrasena incorrecta") from e
    try:
        src.save(str(out_path))
    finally:
        src.close()
    return out_path


# --- Info del PDF -------------------------------------------------------

def pdf_info(pdf_path: Path) -> dict:
    """Devuelve metadata basica del PDF."""
    utils.ensure_pdf(pdf_path)
    doc = fitz.open(str(pdf_path))
    try:
        meta = doc.metadata or {}
        size = pdf_path.stat().st_size
        return {
            "kind": "pdf",
            "pages": doc.page_count,
            "metadata": {k: str(v) for k, v in meta.items() if v},
            "encrypted": doc.is_encrypted,
            "size_bytes": size,
            "size_human": utils.human_size(size),
            "filename": pdf_path.name,
            "outline": [
                {"level": item[0], "title": item[1], "page": item[2]}
                for item in (doc.get_toc() or [])
            ],
        }
    finally:
        doc.close()


# --- Info de documentos (multi-formato) --------------------------------

# Tipos de documento soportados por document_info().
DOC_KINDS = ("pdf", "docx", "xlsx", "pptx", "image", "text")

# Mapeo extension -> kind (Office Open XML y formatos comunes).
_EXT_TO_KIND = {
    ".pdf": "pdf",
    ".docx": "docx",
    ".docm": "docx",
    ".dotx": "docx",
    ".xlsx": "xlsx",
    ".xlsm": "xlsx",
    ".xltx": "xlsx",
    ".pptx": "pptx",
    ".pptm": "pptx",
    ".potx": "pptx",
    ".png": "image", ".jpg": "image", ".jpeg": "image",
    ".webp": "image", ".gif": "image", ".bmp": "image",
    ".tif": "image", ".tiff": "image",
    ".txt": "text", ".md": "text", ".csv": "text", ".log": "text",
}

# Tags EXIF utiles para mostrar en la UI.
_EXIF_LABELS = {
    0x010F: "Fabricante",
    0x0110: "Modelo",
    0x0112: "Orientacion",
    0x011A: "XResolution",
    0x011B: "YResolution",
    0x0131: "Software",
    0x0132: "Fecha",
    0x9003: "Fecha original",
    0x9004: "Fecha digitalizacion",
    0x8825: "GPSInfo",
}


def detect_doc_kind(path: Path) -> str:
    """Detecta el tipo de documento a partir de la extension."""
    return _EXT_TO_KIND.get(path.suffix.lower(), "unknown")


def _common_envelope(path: Path, kind: str) -> Dict[str, Any]:
    """Campos comunes a todos los extractores (nombre, tamano, tipo)."""
    size = path.stat().st_size
    return {
        "kind": kind,
        "filename": path.name,
        "size_bytes": size,
        "size_human": utils.human_size(size),
    }


# --- OOXML: docx, xlsx, pptx (metadata viene en docProps/*.xml) --------

_OOXML_NS_CORE = {
    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dc": "http://purl.org/dc/elements/1.1/",
    "dcterms": "http://purl.org/dc/terms/",
    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
}
_OOXML_NS_APP = {
    "ep": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties",
}


def _read_ooxml_core(path: Path) -> Dict[str, str]:
    """Lee docProps/core.xml (title, author, created, etc.) de un OOXML."""
    out: Dict[str, str] = {}
    try:
        with zipfile.ZipFile(path) as z:
            with z.open("docProps/core.xml") as f:
                import xml.etree.ElementTree as ET
                root = ET.parse(f).getroot()
        for qname, label in (
            ("dc:title", "title"), ("dc:subject", "subject"),
            ("dc:creator", "creator"), ("cp:keywords", "keywords"),
            ("dc:description", "description"), ("cp:lastModifiedBy", "lastModifiedBy"),
            ("cp:revision", "revision"), ("cp:category", "category"),
            ("cp:contentStatus", "contentStatus"),
        ):
            ns_prefix, tag = qname.split(":")
            ns = _OOXML_NS_CORE[ns_prefix]
            el = root.find(f"{{{ns}}}{tag}")
            if el is not None and el.text:
                out[label] = el.text
        # Fechas (vienen como xsd:dateTime en dcterms:created/modified)
        for qname, label in (
            ("dcterms:created", "created"), ("dcterms:modified", "modified"),
        ):
            ns_prefix, tag = qname.split(":")
            ns = _OOXML_NS_CORE[ns_prefix]
            el = root.find(f"{{{ns}}}{tag}")
            if el is not None and el.text:
                out[label] = el.text
    except (KeyError, zipfile.BadZipFile, ET.ParseError):
        pass
    return out


def _read_ooxml_app(path: Path) -> Dict[str, str]:
    """Lee docProps/app.xml (Application, Pages, Words, etc.) de un OOXML."""
    out: Dict[str, str] = {}
    try:
        with zipfile.ZipFile(path) as z:
            with z.open("docProps/app.xml") as f:
                import xml.etree.ElementTree as ET
                root = ET.parse(f).getroot()
        ns = _OOXML_NS_APP["ep"]
        for tag in (
            "Application", "AppVersion", "Template",
            "Company", "Manager", "TotalTime",
            "Pages", "Words", "Characters", "CharactersWithSpaces",
            "Lines", "Paragraphs", "Slides", "Notes",
            "HiddenSlides", "MMClips", "ScaleCrop",
            "PresentationFormat", "DocSecurity",
        ):
            el = root.find(f"{{{ns}}}{tag}")
            if el is not None and el.text:
                out[tag] = el.text
    except (KeyError, zipfile.BadZipFile, ET.ParseError):
        pass
    return out


# --- Word (.docx) -------------------------------------------------------

def _docx_info(path: Path) -> Dict[str, Any]:
    """Metadatos de un .docx. Cuenta tablas, parrafos, palabras, etc."""
    info = _common_envelope(path, "docx")
    core = _read_ooxml_core(path)
    app = _read_ooxml_app(path)
    info["metadata"] = core
    info["app"] = {k: v for k, v in app.items() if k not in ("Template",)}

    # Contar tablas y dar dimensiones leyendo word/document.xml
    tables: List[Dict[str, Any]] = []
    para_count = 0
    word_count = 0
    try:
        with zipfile.ZipFile(path) as z:
            with z.open("word/document.xml") as f:
                import xml.etree.ElementTree as ET
                root = ET.parse(f).getroot()
        ns_w = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
        for i, tbl in enumerate(root.iter(f"{ns_w}tbl"), start=1):
            rows = tbl.findall(f"{ns_w}tr")
            row_count = len(rows)
            col_count = max((len(r.findall(f"{ns_w}tc")) for r in rows), default=0)
            preview = ""
            for r in rows[:2]:
                cell_text = " ".join(
                    (t.text or "") for t in r.iter(f"{ns_w}t") if t.text
                ).strip()
                if cell_text:
                    preview = cell_text[:80]
                    break
            tables.append({
                "index": i,
                "rows": row_count,
                "cols": col_count,
                "preview": preview,
            })
        para_count = sum(1 for _ in root.iter(f"{ns_w}p"))
        word_count = sum(
            len((t.text or "").split())
            for t in root.iter(f"{ns_w}t")
            if t.text
        )
    except (KeyError, zipfile.BadZipFile, ET.ParseError):
        pass
    info["tables"] = tables
    info["stats"] = {
        "paragraphs": para_count,
        "words": word_count,
        "tables": len(tables),
    }
    return info


# --- Excel (.xlsx) ------------------------------------------------------

def _xlsx_info(path: Path) -> Dict[str, Any]:
    """Metadatos de un .xlsx. Lista hojas, sus dimensiones y nombres definidos."""
    info = _common_envelope(path, "xlsx")
    core = _read_ooxml_core(path)
    app = _read_ooxml_app(path)
    info["metadata"] = core
    info["app"] = {k: v for k, v in app.items() if k not in ("Template",)}

    sheets: List[Dict[str, Any]] = []
    total_cells = 0
    defined_names: List[Dict[str, str]] = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                min_col, min_row = ws.min_column, ws.min_row
                max_col, max_row = ws.max_column, ws.max_row
                if max_col is None or max_row is None:
                    sheets.append({"name": ws.title, "rows": 0, "cols": 0,
                                   "cells": 0, "tables": 0, "note": "Hoja vacia"})
                    continue
                rows = max_row - min_row + 1
                cols = max_col - min_col + 1
                # "Tablas" en Excel: usamos el termino "bloques" para no
                # confundir con Excel Tables (que requieren formato especial).
                # Aqui cada hoja se considera un bloque; reportamos dimensiones.
                cells = sum(1 for row in ws.iter_rows() for c in row if c.value is not None)
                total_cells += cells
                sheets.append({
                    "name": ws.title,
                    "rows": rows,
                    "cols": cols,
                    "cells": cells,
                    "tables": 0,  # Excel Tables (ListObjects) requieren inspeccion aparte
                    "state": ws.sheet_state,
                })
            for dn in wb.defined_names:
                try:
                    dests = list(wb.defined_names[dn].destinations) if dn in wb.defined_names else []
                except Exception:
                    dests = []
                defined_names.append({"name": dn, "destinations": str(dests)[:200]})
        finally:
            wb.close()
    except Exception as e:  # noqa: BLE001 - openpyxl lanza varios tipos
        info.setdefault("warnings", []).append(f"No se pudo abrir con openpyxl: {e}")
    info["sheets"] = sheets
    info["defined_names"] = defined_names[:50]  # cap
    info["stats"] = {
        "sheets": len(sheets),
        "cells_with_value": total_cells,
        "defined_names": len(defined_names),
    }
    return info


# --- PowerPoint (.pptx) -------------------------------------------------

def _pptx_info(path: Path) -> Dict[str, Any]:
    """Metadatos de un .pptx. Cuenta slides y tablas por slide."""
    info = _common_envelope(path, "pptx")
    core = _read_ooxml_core(path)
    app = _read_ooxml_app(path)
    info["metadata"] = core
    info["app"] = {k: v for k, v in app.items() if k not in ("Template",)}

    slides: List[Dict[str, Any]] = []
    total_tables = 0
    slide_w = slide_h = None
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            slide_names = sorted(
                [n for n in names if re.match(r"ppt/slides/slide\d+\.xml$", n)],
                key=lambda s: int(re.search(r"slide(\d+)\.xml$", s).group(1)),
            )
            for sn in slide_names:
                idx = int(re.search(r"slide(\d+)\.xml$", sn).group(1))
                with z.open(sn) as f:
                    import xml.etree.ElementTree as ET
                    root = ET.parse(f).getroot()
                ns_a = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
                tables = list(root.iter(f"{ns_a}tbl"))
                total_tables += len(tables)
                table_dims = []
                for t in tables:
                    rows = t.findall(f"{ns_a}tr")
                    cols = max((len(r.findall(f"{ns_a}tc")) for r in rows), default=0)
                    table_dims.append({"rows": len(rows), "cols": cols})
                slides.append({
                    "index": idx,
                    "tables": len(tables),
                    "table_dims": table_dims,
                })
            # Tamano de la presentacion (EMU -> cm: /360000)
            if "ppt/presentation.xml" in names:
                with z.open("ppt/presentation.xml") as f:
                    import xml.etree.ElementTree as ET
                    root = ET.parse(f).getroot()
                ns_p = "{http://schemas.openxmlformats.org/presentationml/2006/main}"
                sld = root.find(f"{ns_p}sldSz")
                if sld is not None:
                    cx = int(sld.get("cx", 0))
                    cy = int(sld.get("cy", 0))
                    if cx and cy:
                        slide_w = round(cx / 360000, 2)
                        slide_h = round(cy / 360000, 2)
    except (KeyError, zipfile.BadZipFile, ET.ParseError):
        pass
    info["slides"] = slides
    info["stats"] = {
        "slides": len(slides),
        "tables": total_tables,
        "slide_w_cm": slide_w,
        "slide_h_cm": slide_h,
    }
    return info


# --- Imagenes (png, jpg, webp, etc.) ------------------------------------

def _image_info(path: Path) -> Dict[str, Any]:
    """Metadatos de una imagen: tamano, formato, modo, EXIF basico."""
    info = _common_envelope(path, "image")
    try:
        with Image.open(path) as im:
            info["format"] = im.format
            info["mode"] = im.mode
            info["width"] = im.width
            info["height"] = im.height
            info["megapixels"] = round((im.width * im.height) / 1_000_000, 2)
            dpi = im.info.get("dpi")
            if dpi:
                info["dpi"] = f"{int(dpi[0])} x {int(dpi[1])}" if len(dpi) == 2 else str(dpi)
            # EXIF
            exif = {}
            try:
                raw = im.getexif()
                if raw:
                    for tag_id, val in raw.items():
                        label = _EXIF_LABELS.get(tag_id)
                        if not label:
                            continue
                        # Decodifica bytes para legibilidad
                        if isinstance(val, bytes):
                            try:
                                val = val.decode("utf-8", errors="replace").strip("\x00").strip()
                            except Exception:
                                val = repr(val)
                        exif[label] = str(val)[:200]
            except Exception:
                pass
            if exif:
                info["exif"] = exif
            info["stats"] = {
                "ancho": im.width,
                "alto": im.height,
                "megapixeles": info["megapixels"],
                "modo": im.mode,
            }
    except Exception as e:  # noqa: BLE001
        info["error"] = f"No se pudo abrir la imagen: {e}"
    return info


# --- Texto plano --------------------------------------------------------

def _text_info(path: Path) -> Dict[str, Any]:
    """Metadatos de un .txt/.md/.csv: encoding, lineas, palabras, caracteres."""
    info = _common_envelope(path, "text")
    raw = path.read_bytes()
    # Intentar decodificar con UTF-8 primero, fallback a latin-1 (nunca falla)
    encoding = "utf-8"
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
        encoding = "utf-8 (con reemplazos)"
    lines = text.count("\n") + (0 if text.endswith("\n") else 1)
    words = len(text.split())
    chars = len(text)
    chars_no_ws = len(re.sub(r"\s", "", text))
    # Para CSV: contar columnas (separador mas comun)
    extra = {}
    if path.suffix.lower() == ".csv":
        sample = text[:4096]
        for sep in (",", ";", "\t", "|"):
            counts = [line.count(sep) for line in sample.splitlines()[:10] if line]
            if counts and max(counts) == min(counts) and max(counts) > 0:
                extra["separador"] = sep
                extra["columnas"] = max(counts) + 1
                break
    info["encoding"] = encoding
    info["stats"] = {
        "lineas": lines,
        "palabras": words,
        "caracteres": chars,
        "caracteres_sin_espacios": chars_no_ws,
    }
    if extra:
        info["stats"].update(extra)
    return info


# --- Dispatcher ---------------------------------------------------------

def document_info(path: Path) -> Dict[str, Any]:
    """
    Punto de entrada unico: detecta el tipo de archivo y devuelve su
    metadata normalizada en un dict. La UI consume el campo 'kind' para
    decidir como renderizar.
    """
    if not path.exists():
        raise FileNotFoundError(f"No existe: {path}")
    if path.stat().st_size == 0:
        raise ValueError("Archivo vacio")
    kind = detect_doc_kind(path)
    if kind == "pdf":
        return pdf_info(path)
    if kind == "docx":
        return _docx_info(path)
    if kind == "xlsx":
        return _xlsx_info(path)
    if kind == "pptx":
        return _pptx_info(path)
    if kind == "image":
        return _image_info(path)
    if kind == "text":
        return _text_info(path)
    raise ValueError(
        f"Formato no soportado: {path.suffix or '(sin extension)'}. "
        f"Soportados: PDF, Word (.docx), Excel (.xlsx), "
        f"PowerPoint (.pptx), imagenes, texto."
    )
